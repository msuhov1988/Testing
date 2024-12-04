"""
Модуль для чтения исходных данных из xlsx файла
Файл определяется по заданному имени - testing_data, по заданному пути - корневая папка проекта
"""

import os.path
from enum import Enum
from collections import namedtuple, defaultdict
from datetime import datetime, date
import zipfile
from typing import Union, Tuple, Dict, List, TypeAlias
import openpyxl
from app.constants import DATE_BASEMENT


# необходимые колонки исходного файла
class _Columns(Enum):
    state = "Состояние заявки"
    status = "Статус заявки"
    author = "Автор заявки"
    creation_dt = "Дата создания заявки"
    package_id = "ID пакета"


_ColNameIndexMapping = namedtuple("_ColNameIndexMapping",
                                  ("state", "status", "author", "creation_dt", "package_id"))


def _check_columns(worksheet) -> Union[Tuple[int, _ColNameIndexMapping], None]:
    for row_ind, row in enumerate(worksheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
        mapping = dict()
        for col_ind, cell in enumerate(row):
            if cell == _Columns.state.value:
                mapping[_Columns.state.value] = col_ind
            elif cell == _Columns.status.value:
                mapping[_Columns.status.value] = col_ind
            elif cell == _Columns.author.value:
                mapping[_Columns.author.value] = col_ind
            elif cell == _Columns.creation_dt.value:
                mapping[_Columns.creation_dt.value] = col_ind
            elif cell == _Columns.package_id.value:
                mapping[_Columns.package_id.value] = col_ind
        if len(mapping.keys()) == len(_Columns):
            return row_ind + 1, _ColNameIndexMapping(*(mapping[col.value] for col in _Columns))
    return None


# необходимые значения столбца Columns.state.value
class _StateKeys(Enum):
    double_part = "Дубликат заявки"
    for_creation = "ДОБАВЛЕНИЕ"
    for_expand = "РАСШИРЕНИЕ"


# необходимые значения столбца Columns.status.value
class _StatusKeys(Enum):
    processing_over = "Обработка завершена"
    returned = "Возвращена на уточнение"
    sent_for_processing = "Отправлена в обработку"


class _DailyData:
    def __init__(self):
        self.loaded: int = 0
        self.doubles: int = 0
        self.for_creation: int = 0
        self.for_expand: int = 0
        self.handle_over: int = 0
        self.returned: int = 0
        self.sent_for_handle: int = 0

        self.packages = set()
        self.users = set()

    def add_data(self, state: str, status: str, author: str, package_id: str):
        self.loaded += 1
        if _StateKeys.double_part.value in state:
            self.doubles += 1
        elif state == _StateKeys.for_creation.value:
            self.for_creation += 1
        elif state == _StateKeys.for_expand.value:
            self.for_expand += 1
        if status == _StatusKeys.processing_over.value:
            self.handle_over += 1
        elif status == _StatusKeys.returned.value:
            self.returned += 1
        elif status == _StatusKeys.sent_for_processing.value:
            self.sent_for_handle += 1

        self.packages.add(package_id)
        self.users.add(author)

    def output(self):
        requests_qnt = (self.loaded, self.doubles, self.for_creation, self.for_expand,
                        self.handle_over, self.returned, self.sent_for_handle, len(self.packages))
        users = tuple(self.users)
        return requests_qnt, users


def _collect(worksheet, start_row: int,
             col_mapping: _ColNameIndexMapping) -> Union[Tuple[Dict[date, _DailyData], None], Tuple[None, str]]:
    result = defaultdict(_DailyData)
    for row_ind, row in enumerate(worksheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
        state, status, author = row[col_mapping.state], row[col_mapping.status], row[col_mapping.author]
        creation, package_id = row[col_mapping.creation_dt], row[col_mapping.package_id]
        if isinstance(creation, datetime):
            creation_dt = creation.date()
        else:
            try:
                creation_dt = datetime.strptime(creation, "%d.%m.%Y %H:%M:%S").date()
            except ValueError:
                return None, f"Строка {row_ind}, Столбец {col_mapping.creation_dt + 1} - неверный формат строки-даты"

        daily_data = result[creation_dt]
        daily_data.add_data(state=state, status=status, author=author, package_id=package_id)
    return result, None


_TransformedData: TypeAlias = Tuple[int, int, List[Tuple], List[Tuple]]


def _transform(data: Dict[date, _DailyData]) -> _TransformedData:
    min_date_int, max_date_int = 0, 0
    requests_qnt, users = list(), list()
    for dt, daily_data in data.items():
        date_to_int = (dt - DATE_BASEMENT).days
        if min_date_int == 0 or date_to_int < min_date_int:
            min_date_int = date_to_int
        if max_date_int == 0 or date_to_int > max_date_int:
            max_date_int = date_to_int

        daily_qnt, daily_users = daily_data.output()
        users.extend(((date_to_int, user) for user in daily_users))
        requests_qnt.append((date_to_int, *daily_qnt))

    return min_date_int, max_date_int, users, requests_qnt


_SOURCE_FILE = os.path.join(os.path.dirname(__file__), "testing_data.xlsx")
_SHEET_WITH_DATA = "Data"


def read_data_from_file() -> Union[Tuple[_TransformedData, None], Tuple[None, str]]:
    """
    Единственная импортируемая функция модуля. Открывает и читает данные их xlsx файла
    Ищет необходимые колонки по их наименованию, порядок следования не важен
    :return: кортеж (результат, ошибка)
    """
    wb = None
    try:
        wb = openpyxl.load_workbook(filename=_SOURCE_FILE, read_only=True, data_only=True)
        ws = wb[_SHEET_WITH_DATA] if _SHEET_WITH_DATA in wb.sheetnames else wb.active
        check = _check_columns(worksheet=ws)
        if check is None:
            return None, f"!ОШИБКА ФАЙЛА - {_SOURCE_FILE} - названия колонок в файле не соответствуют требуемым"
        start_row, col_mapping = check
        collect, error = _collect(worksheet=ws, start_row=start_row, col_mapping=col_mapping)
        if error is not None:
            return None, f"!ОШИБКА ФАЙЛА - {error}"
        transform = _transform(data=collect)
        return transform, None
    except zipfile.BadZipfile:
        return None, f"!ОШИБКА ФАЙЛА - {_SOURCE_FILE} - не является файлом формата xlsx"
    except FileNotFoundError:
        return None, f"!ОШИБКА ФАЙЛА - {_SOURCE_FILE} - не существует"
    finally:
        if wb is not None:
            wb.close()
